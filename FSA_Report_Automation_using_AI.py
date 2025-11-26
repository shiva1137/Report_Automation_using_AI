import asyncio
import calendar
import gc
import json
import logging
import os
import re
import sys
import threading
import time
from datetime import datetime, timedelta
from logging.handlers import TimedRotatingFileHandler
from typing import Dict, Optional, Tuple

# Try to load python-dotenv for .env file support (optional)
# If not installed, environment variables can be set manually
try:
    from dotenv import load_dotenv
    load_dotenv()  # Load environment variables from .env file if it exists
except ImportError:
    # python-dotenv is optional - environment variables can be set manually
    # Install with: pip install python-dotenv
    pass

import pandas as pd
import pytz
import requests
from motor.motor_asyncio import AsyncIOMotorClient
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment
from pymongo.errors import PyMongoError
from tenacity import retry, stop_after_attempt, wait_exponential, retry_if_exception_type
from telegram import Update
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes, ConversationHandler
from openai import OpenAI


class Config:
    """Configuration settings for the script.
    
    All sensitive credentials are loaded from environment variables for security.
    See .env.example for required environment variables.
    """
    # MongoDB Configuration - Load from environment variable
    MONGO_CONNECTION_STRING = os.getenv(
        "MONGO_CONNECTION_STRING",
        ""  # Default empty - must be set via environment variable
    )
    
    # Telegram Bot Configuration - Load from environment variables
    TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN", "")
    
    # Telegram Chat IDs - Comma-separated list, converted to integers
    _telegram_chat_ids_str = os.getenv("TELEGRAM_CHAT_ID", "")
    if _telegram_chat_ids_str:
        TELEGRAM_CHAT_ID = [int(cid.strip()) for cid in _telegram_chat_ids_str.split(",") if cid.strip()]
    else:
        TELEGRAM_CHAT_ID = []  # Default empty - must be set via environment variable
    
    # LLM7.io API Configuration
    LLM7_API_KEY = os.getenv("LLM7_API_KEY", "")
    LLM7_BASE_URL = os.getenv("LLM7_BASE_URL", "https://api.llm7.io/v1")
    LLM7_MODEL = os.getenv("LLM7_MODEL", "gpt-4o")  # GPT-4 Omni - best model for structured output
    
    # Application Settings
    SCRIPT_NAME = os.path.splitext(os.path.basename(__file__))[0]
    LOG_FILE_NAME = f"{SCRIPT_NAME}_Log.log"
    MAX_WORKERS = int(os.getenv("MAX_WORKERS", "500"))
    MONGO_IDLE_TIMEOUT = int(os.getenv("MONGO_IDLE_TIMEOUT", "300"))  # 5 minutes default
    
    # Business Logic Configuration
    CATEGORIES = ["MC", "JR", "PS", "DFW"]  # Trip categories
    
    # Timezone configuration
    TIMEZONE = pytz.timezone(os.getenv("TIMEZONE", "Asia/Kolkata"))
    
    # Available areas for trip data (can be customized via environment variable)
    # Format: Comma-separated list, or use default
    _areas_str = os.getenv("AREAS", "")
    if _areas_str:
        AREAS = [area.strip() for area in _areas_str.split(",") if area.strip()]
    else:
        # Default areas (can be customized for your use case)
        AREAS = [
            "01-Thiruvottiyur(Area-1)", "02-Manali(Area-2)", "03-Madhavaram(Area-3)", 
            "04-Tondiarpet(Area-4)", "05-Royapuram(Area-5)", "06-Thiru-Vi-Ka Nagar(Area-6)", 
            "07-Ambattur(Area-7)", "08-Anna Nagar(Area-8)", "09-Teynampet(Area-9)", 
            "10-Kodambakkam(Area-10)", "11-Valasaravakkam(Area-11)", "12-Alandur(Area-12)",
            "13-Adyar(Area-13)", "14-Perungudi(Area-14)", "15-Sholinganallur(Area-15)"
        ]
    
    @classmethod
    def validate(cls) -> bool:
        """Validate that all required configuration is present."""
        missing = []
        if not cls.MONGO_CONNECTION_STRING:
            missing.append("MONGO_CONNECTION_STRING")
        if not cls.TELEGRAM_BOT_TOKEN:
            missing.append("TELEGRAM_BOT_TOKEN")
        if not cls.TELEGRAM_CHAT_ID:
            missing.append("TELEGRAM_CHAT_ID")
        if not cls.LLM7_API_KEY:
            missing.append("LLM7_API_KEY")
        
        if missing:
            print("❌ Missing required environment variables:")
            for var in missing:
                print(f"   - {var}")
            print("\n📝 Please set these in your .env file or environment variables.")
            print("   See .env.example for reference.")
            return False
        return True

# Conversation states
WAITING_FOR_PERIOD, WAITING_FOR_AREA = range(2)


class MongoConnectionManager:
    """Simplified MongoDB connection manager - tracks last message time and auto-closes after idle timeout."""
    _instance = None
    _lock = threading.Lock()
    
    def __new__(cls):
        if cls._instance is None:
            with cls._lock:
                if cls._instance is None:
                    cls._instance = super(MongoConnectionManager, cls).__new__(cls)
                    cls._instance._initialized = False
        return cls._instance
    
    def __init__(self):
        if self._initialized:
            return
        
        self.client: Optional[AsyncIOMotorClient] = None
        self.last_message_time: float = 0  # Track last message time from user
        self.logger: Optional[logging.Logger] = None
        self._cleanup_task: Optional[asyncio.Task] = None
        self._lock = asyncio.Lock()
        self._initialized = True
    
    async def _check_and_close_idle(self):
        """Check every 1 minute if last message time > 5 minutes, then close connection."""
        try:
            while True:
                await asyncio.sleep(60)  # Check every 1 minute
                
                async with self._lock:
                    if self.client is None:
                        break  # Exit if connection already closed
                    
                    current_time = time.time()
                    time_since_last_message = current_time - self.last_message_time
                    
                    # If last message was more than 5 minutes ago, close connection
                    if time_since_last_message >= Config.MONGO_IDLE_TIMEOUT:
                        await self.close()
                        break  # Exit task
        except asyncio.CancelledError:
            await self.close(force=True)
        except Exception as e:
            if self.logger:
                self.logger.error(f"Error in cleanup task: {str(e)}")
    
    async def get_client(self, logger: logging.Logger) -> AsyncIOMotorClient:
        """Get or create MongoDB client. Updates last message time when user sends message."""
        async with self._lock:
            # Update last message time when getting client (called when user sends message)
            self.last_message_time = time.time()
            
            # Create new client if None
            if self.client is None:
                self.logger = logger
                self.client = AsyncIOMotorClient(
                    Config.MONGO_CONNECTION_STRING,
                    maxPoolSize=Config.MAX_WORKERS,
                    serverSelectionTimeoutMS=5000
                )
                logger.info("Created new MongoDB connection")
                
                # Start background cleanup task if not running
                if self._cleanup_task is None or self._cleanup_task.done():
                    self._cleanup_task = asyncio.create_task(self._check_and_close_idle())
            
            return self.client
    
    async def close(self, force: bool = False):
        """Close MongoDB connection."""
        async with self._lock:
            if self.client is None:
                return
            
            try:
                self.client.close()
                idle_time = time.time() - self.last_message_time if self.last_message_time > 0 else 0
                if self.logger:
                    self.logger.info(f"MongoDB connection closed (idle: {idle_time:.1f}s, forced: {force})")
            except Exception as e:
                if self.logger:
                    self.logger.error(f"Error closing MongoDB connection: {str(e)}")
            finally:
                self.client = None
                gc.collect()  # Trigger garbage collection

# Global connection manager instance
_mongo_manager = MongoConnectionManager()


def setup_logger(current_dir: str) -> logging.Logger:
    """Configure logging with file and console handlers."""
    logger = logging.getLogger(__name__)
    logger.setLevel(logging.INFO)

    if not logger.handlers:
        log_path = os.path.join(current_dir, Config.LOG_FILE_NAME)
        file_handler = TimedRotatingFileHandler(log_path, when='midnight', interval=1, backupCount=5)
        console_handler = logging.StreamHandler()

        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
        file_handler.setFormatter(formatter)
        console_handler.setFormatter(formatter)

        logger.addHandler(file_handler)
        logger.addHandler(console_handler)

    return logger


def get_openai_client(logger: logging.Logger):
    """Get OpenAI-compatible client configured for llm7.io ONLY."""
    try:
        if not Config.LLM7_API_KEY:
            logger.error("LLM7_API_KEY not set in environment variables")
            logger.info("Please set LLM7_API_KEY environment variable with your llm7.io API key")
            return None
        # Create client with llm7.io endpoint
        client = OpenAI(
            api_key=Config.LLM7_API_KEY,
            base_url=Config.LLM7_BASE_URL
        )
        logger.debug(f"Using LLM7.io endpoint with model: {Config.LLM7_MODEL}")
        return client
    except Exception as e:
        logger.error(f"Error creating LLM7.io client: {str(e)}")
        return None


def parse_date_from_text(text: str, logger: logging.Logger) -> Optional[Tuple[datetime, datetime]]:
    """Parse date/period from text using NLP and return (start_date, end_date) tuple.
    
    Supports various date formats including:
    - "Jun 2024", "June 2024", "Jun-2024"
    - "Jun 2023", "June 2023"
    - "Jan 2025", "January 2025"
    - "2024", "2025" (full year)
    - Date ranges: "Jun 2024 to Aug 2024", "June 2024 to August 2024"
    - Month only (finds last occurrence): "August", "Aug" → Last August month
    """
    try:
        client = get_openai_client(logger)
        if not client:
            return None
        
        current_date = datetime.now(Config.TIMEZONE)
        current_year = current_date.year
        current_month = current_date.month
        
        # Check if it's a month-only query (no year mentioned)
        month_only_pattern = r'\b(?:Jan|January|Feb|February|Mar|March|Apr|April|May|Jun|June|Jul|July|Aug|August|Sep|September|Oct|October|Nov|November|Dec|December)\b'
        year_pattern = r'\b(?:19|20)\d{2}\b'
        
        has_year = bool(re.search(year_pattern, text, re.IGNORECASE))
        has_month = bool(re.search(month_only_pattern, text, re.IGNORECASE))
        
        # If month-only query, find last occurrence of that month
        if has_month and not has_year:
            month_names_map = {
                'jan': 1, 'january': 1, 'feb': 2, 'february': 2, 'mar': 3, 'march': 3,
                'apr': 4, 'april': 4, 'may': 5, 'jun': 6, 'june': 6, 'jul': 7, 'july': 7,
                'aug': 8, 'august': 8, 'sep': 9, 'september': 9, 'oct': 10, 'october': 10,
                'nov': 11, 'november': 11, 'dec': 12, 'december': 12
            }
            
            text_lower = text.lower()
            matched_month = None
            for month_name, month_num in month_names_map.items():
                if month_name in text_lower:
                    matched_month = month_num
                    break
            
            if matched_month:
                # Find last occurrence of this month (before current date)
                target_year = current_year
                if current_month < matched_month:
                    # Last occurrence was in previous year
                    target_year = current_year - 1
                elif current_month == matched_month:
                    # Current month, use it
                    target_year = current_year
                else:
                    # Last occurrence was earlier this year
                    target_year = current_year
                
                # Get first and last day of that month
                first_day = datetime(target_year, matched_month, 1).date()
                last_day_num = calendar.monthrange(target_year, matched_month)[1]
                last_day = datetime(target_year, matched_month, last_day_num).date()
                
                start_time = Config.TIMEZONE.localize(datetime.combine(first_day, datetime.min.time()))
                end_time = Config.TIMEZONE.localize(datetime.combine(last_day, datetime.max.time()))
                
                logger.info(f"Month-only query detected: {text} → Last occurrence: {first_day} to {last_day}")
                return (start_time, end_time)
        
        # Get model name
        model_name = Config.LLM7_MODEL
        
        # Enhanced prompt with comprehensive date format support including date ranges
        prompt = f"""Extract the date/period from the following user query. 

IMPORTANT: Support ALL common date formats including:
- Month abbreviations: "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"
- Full month names: "January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"
- Month Year: "Jun 2024", "June 2024", "Jun-2024", "Jun/2024"
- Year only: "2024", "2025"
- DATE RANGES: "Jun 2024 to Aug 2024", "June 2024 to August 2024", "Jun 2024 - Aug 2024"
- Past dates: "Jun 2023", "June 2023", "2023"

Current date: {current_date.strftime('%Y-%m-%d')} ({current_date.strftime('%B %Y')})

User query: "{text}"

Return ONLY a JSON object with this exact structure:
{{
    "start_date": "YYYY-MM-DD",
    "end_date": "YYYY-MM-DD"
}}

Examples:
- "Jun 2024" → {{"start_date": "2024-06-01", "end_date": "2024-06-30"}}
- "June 2024" → {{"start_date": "2024-06-01", "end_date": "2024-06-30"}}
- "Jun 2023" → {{"start_date": "2023-06-01", "end_date": "2023-06-30"}}
- "Jun 2024 to Aug 2024" → {{"start_date": "2024-06-01", "end_date": "2024-08-31"}}
- "June 2024 to August 2024" → {{"start_date": "2024-06-01", "end_date": "2024-08-31"}}
- "Jan 2025" → {{"start_date": "2025-01-01", "end_date": "2025-01-31"}}
- "2024" → {{"start_date": "2024-01-01", "end_date": "2024-12-31"}}
- "2025" → {{"start_date": "2025-01-01", "end_date": "2025-12-31"}}

For a single month, return the first and last day of that month.
For a DATE RANGE like "Jun 2024 to Aug 2024", return start of first month to end of last month.
For a full year, return January 1 to December 31 of that year.

If no date is found, return null for both dates.
"""
        
        # Try with JSON format first, fallback to regular if not supported
        try:
            response = client.chat.completions.create(
                model=model_name,
                messages=[
                    {"role": "system", "content": "You are an expert date parsing assistant. Extract dates from user queries accurately and return only valid JSON. Handle all month abbreviations and formats correctly."},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.1,
                max_tokens=150,
                response_format={"type": "json_object"}  # Request JSON format explicitly
            )
        except Exception as e:
            logger.warning(f"JSON format not supported, trying without: {str(e)}")
            # Fallback: request JSON in system message
            response = client.chat.completions.create(
                model=model_name,
                messages=[
                    {"role": "system", "content": "You are an expert date parsing assistant. Extract dates from user queries accurately and return ONLY valid JSON. Handle all month abbreviations and formats correctly. Return JSON only, no markdown or explanations."},
                    {"role": "user", "content": prompt + "\n\nIMPORTANT: Return ONLY valid JSON, no markdown code blocks or explanations."}
                ],
                temperature=0.1,
                max_tokens=150
            )
        
        result_text = response.choices[0].message.content.strip()
        
        # Clean up the response (remove markdown code blocks if present)
        result_text = re.sub(r'```json\s*', '', result_text)
        result_text = re.sub(r'```\s*', '', result_text)
        result_text = result_text.strip()
        
        # Parse JSON response
        result = json.loads(result_text)
        
        if result.get("start_date") and result.get("end_date"):
            try:
                start_date = datetime.strptime(result["start_date"], "%Y-%m-%d").date()
                end_date = datetime.strptime(result["end_date"], "%Y-%m-%d").date()
                
                # Validate date range
                if start_date > end_date:
                    logger.warning(f"Invalid date range: start_date ({start_date}) > end_date ({end_date})")
                    return None
                
                # Convert to timezone-aware datetime
                start_time = Config.TIMEZONE.localize(datetime.combine(start_date, datetime.min.time()))
                end_time = Config.TIMEZONE.localize(datetime.combine(end_date, datetime.max.time()))
                
                logger.info(f"Successfully parsed date: {start_date} to {end_date} from query: '{text}'")
                return (start_time, end_time)
            except ValueError as e:
                logger.error(f"Error parsing date format: {str(e)}")
                return None
        
        logger.warning(f"No valid date found in query: '{text}'")
        return None
        
    except json.JSONDecodeError as e:
        logger.error(f"Error parsing JSON response for date: {str(e)}. Response: {result_text[:200]}")
        return None
    except Exception as e:
        logger.error(f"Error parsing date from text '{text}': {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return None


def parse_query_with_nlp(query: str, logger: logging.Logger) -> Dict[str, Optional[str]]:
    """Parse user query using LLM7.io GPT-4o to extract category, area, and period.
    
    Extracts structured information including:
    - Categories: Can be single (PS) or multiple (PS, MC, JR, DFW) or "all"
    - Areas: Can be single area or multiple areas (Area 1 and Area 2) or "all"
    - Period: Date/period text (e.g., "Jun 2024", "June 2023", "Jun 2024 to Aug 2024", "August")
    """
    try:
        client = get_openai_client(logger)
        if not client:
            logger.error("Failed to create LLM7.io client")
            return {"categories": [], "category": None, "areas": [], "area": None, "period": None, "has_period": False, "has_area": False, "all_categories": False, "all_areas": False}
        
        # Get model name
        model_name = Config.LLM7_MODEL
        
        # Create area mapping for better understanding
        area_mapping = {f"Area-{i}": area for i, area in enumerate(Config.AREAS, 1)}
        area_list = "\n".join([f"- {area}" for area in Config.AREAS])
        
        # Enhanced prompt supporting multiple categories, multiple areas, and date ranges
        prompt = f"""Extract structured information from the following user query about generating Excel files for trip data.

Available Categories: {', '.join(Config.CATEGORIES)}

IMPORTANT: Users may request:
- Single category: "PS trips", "MC trips"
- Multiple categories: "PS and MC trips", "PS, MC, JR trips"
- All categories: "all categories", "all trips", "all"

Available Areas:
{area_list}

IMPORTANT: Users may request:
- Single area: "Area 1", "Area-1", "01-Thiruvottiyur(Area-1)"
- Multiple areas: "Area 1 and Area 2", "Area-1, Area-2", "Area 1, 2, 3"
- All areas: "all areas", "all" (for areas)

Area aliases (user might say):
- "Area-1" or "Area 1" or "Area -1" → 01-Thiruvottiyur(Area-1)
- "Area-2" or "Area 2" → 02-Manali(Area-2)
- "Area-5" or "Area 5" → 05-Royapuram(Area-5)
- "all areas" or "all" (for areas) → ["all"]
- And so on...

Date/Period formats to recognize (extract the EXACT text from query):
- "Jun 2024", "June 2024", "Jun-2024"
- "Jun 2023", "June 2023"
- "Jan 2025", "January 2025"
- "Jun 2024 to Aug 2024", "June 2024 to August 2024" (DATE RANGES)
- "2024", "2025" (full year)
- "August", "Aug" (month only - find last occurrence)
- Any date format mentioned in the query

User query: "{query}"

Return ONLY a JSON object with this exact structure:
{{
    "categories": ["PS", "MC"] or ["all"] or ["PS"] or [],
    "category": "PS|MC|JR|DFW or null" (for backward compatibility, first category if multiple),
    "areas": ["01-Thiruvottiyur(Area-1)", "02-Manali(Area-2)"] or ["all"] or ["01-Thiruvottiyur(Area-1)"] or [],
    "area": "exact area name from available areas list or null" (for backward compatibility, first area if multiple),
    "period": "extracted period text exactly as written in query or null",
    "has_period": true/false,
    "has_area": true/false,
    "all_categories": true/false (true if user wants all categories),
    "all_areas": true/false (true if user wants all areas)
}}

Examples:
- "Give me Excel file for PS trips for Area -1 for Jun 2024"
  → {{"categories": ["PS"], "category": "PS", "areas": ["01-Thiruvottiyur(Area-1)"], "area": "01-Thiruvottiyur(Area-1)", "period": "Jun 2024", "has_period": true, "has_area": true, "all_categories": false, "all_areas": false}}

- "Give me Excel for PS and MC trips for Area -1 for Jun 2024"
  → {{"categories": ["PS", "MC"], "category": "PS", "areas": ["01-Thiruvottiyur(Area-1)"], "area": "01-Thiruvottiyur(Area-1)", "period": "Jun 2024", "has_period": true, "has_area": true, "all_categories": false, "all_areas": false}}

- "PS trips for Area 1 and Area 2 for Jun 2024"
  → {{"categories": ["PS"], "category": "PS", "areas": ["01-Thiruvottiyur(Area-1)", "02-Manali(Area-2)"], "area": "01-Thiruvottiyur(Area-1)", "period": "Jun 2024", "has_period": true, "has_area": true, "all_categories": false, "all_areas": false}}

- "All areas for Jun 2024"
  → {{"categories": ["all"], "category": null, "areas": ["all"], "area": null, "period": "Jun 2024", "has_period": true, "has_area": true, "all_categories": true, "all_areas": true}}

- "Get Excel for all categories for Area 1 Jun 2024 to Aug 2024"
  → {{"categories": ["all"], "category": null, "areas": ["01-Thiruvottiyur(Area-1)"], "area": "01-Thiruvottiyur(Area-1)", "period": "Jun 2024 to Aug 2024", "has_period": true, "has_area": true, "all_categories": true, "all_areas": false}}

- "August trips"
  → {{"categories": ["all"], "category": null, "areas": ["all"], "area": null, "period": "August", "has_period": true, "has_area": false, "all_categories": true, "all_areas": true}}

- "PS, MC, JR trips Area 5 for June 2023"
  → {{"categories": ["PS", "MC", "JR"], "category": "PS", "areas": ["05-Royapuram(Area-5)"], "area": "05-Royapuram(Area-5)", "period": "June 2023", "has_period": true, "has_area": true, "all_categories": false, "all_areas": false}}

- "MC trips Area 5 for June 2024 to August 2024"
  → {{"categories": ["MC"], "category": "MC", "areas": ["05-Royapuram(Area-5)"], "area": "05-Royapuram(Area-5)", "period": "June 2024 to August 2024", "has_period": true, "has_area": true, "all_categories": false, "all_areas": false}}

Extract:
- categories: Array of category codes found, or ["all"] if user wants all categories, or [] if none found
- category: First category for backward compatibility (can be null if all_categories is true)
- areas: Array of exact area names from list, or ["all"] if user wants all areas, or [] if none found
- area: First area for backward compatibility (can be null if all_areas is true)
- period: Exact period text from query, or null
- has_period, has_area: boolean flags
- all_categories: true if user wants all categories
- all_areas: true if user wants all areas
"""
        
        # Try with JSON format first, fallback to regular if not supported
        try:
            response = client.chat.completions.create(
                model=model_name,
                messages=[
                    {"role": "system", "content": "You are an expert query parser assistant. Extract structured information from user queries accurately and return only valid JSON. Match areas exactly to the provided list."},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.1,
                max_tokens=250,
                response_format={"type": "json_object"}  # Request JSON format explicitly
            )
        except Exception as e:
            logger.warning(f"JSON format not supported, trying without: {str(e)}")
            # Fallback: request JSON in system message
            response = client.chat.completions.create(
                model=model_name,
                messages=[
                    {"role": "system", "content": "You are an expert query parser assistant. Extract structured information from user queries accurately and return ONLY valid JSON. Match areas exactly to the provided list. Return JSON only, no markdown or explanations."},
                    {"role": "user", "content": prompt + "\n\nIMPORTANT: Return ONLY valid JSON, no markdown code blocks or explanations."}
                ],
                temperature=0.1,
                max_tokens=250
            )
        
        result_text = response.choices[0].message.content.strip()
        
        # Clean up the response
        result_text = re.sub(r'```json\s*', '', result_text)
        result_text = re.sub(r'```\s*', '', result_text)
        result_text = result_text.strip()
        
        result = json.loads(result_text)
        
        logger.info(f"Parsed query result: {result}")
        
        # Validate result structure
        if not isinstance(result, dict):
            logger.error(f"Invalid result structure: {result}")
            return {"category": None, "area": None, "period": None, "has_period": False, "has_area": False}
        
        # Ensure all required keys exist
        result.setdefault("categories", [])
        result.setdefault("category", None)
        result.setdefault("areas", [])
        result.setdefault("area", None)
        result.setdefault("period", None)
        result.setdefault("has_period", False)
        result.setdefault("has_area", False)
        result.setdefault("all_categories", False)
        result.setdefault("all_areas", False)
        
        # Handle backward compatibility - if category exists but categories doesn't
        if result.get("category") and not result.get("categories"):
            result["categories"] = [result["category"]]
        
        # Handle backward compatibility - if area exists but areas doesn't
        if result.get("area") and not result.get("areas"):
            result["areas"] = [result["area"]]
        
        # Handle "all" in categories
        if result.get("all_categories") or (result.get("categories") and "all" in result["categories"]):
            result["all_categories"] = True
            result["categories"] = ["all"]
        
        # Handle "all" in areas
        if result.get("all_areas") or (result.get("areas") and "all" in result["areas"]):
            result["all_areas"] = True
            result["areas"] = ["all"]
        
        return result
        
    except json.JSONDecodeError as e:
        logger.error(f"Error parsing JSON response: {str(e)}. Response: {result_text[:200] if 'result_text' in locals() else 'N/A'}")
        return {"categories": [], "category": None, "areas": [], "area": None, "period": None, "has_period": False, "has_area": False, "all_categories": False, "all_areas": False}
    except Exception as e:
        logger.error(f"Error parsing query with NLP '{query}': {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return {"categories": [], "category": None, "areas": [], "area": None, "period": None, "has_period": False, "has_area": False, "all_categories": False, "all_areas": False}


def sanitize_filename(text: str) -> str:
    """Sanitize area name for use in filename."""
    # Replace special characters with underscores
    text = re.sub(r'[<>:"/\\|?*]', '_', text)
    # Replace spaces with underscores
    text = text.replace(' ', '_')
    # Remove parentheses and their contents, but keep the area number
    text = re.sub(r'\([^)]*\)', '', text)
    # Clean up multiple underscores
    text = re.sub(r'_+', '_', text)
    # Remove leading/trailing underscores
    text = text.strip('_')
    return text




@retry(stop=stop_after_attempt(3), wait=wait_exponential(multiplier=1, min=4, max=10), retry=retry_if_exception_type(requests.RequestException))
async def send_to_telegram(file_path: str, logger: logging.Logger, area: str, current_category: str, month_year: str, trip_count: int, chat_id: int = None) -> None:
    """Send message with Excel file to Telegram."""
    try:
        # Use chat_id if provided and authorized, otherwise use first authorized chat ID
        if chat_id is None:
            chat_id = Config.TELEGRAM_CHAT_ID[0]  # Use first authorized chat ID if not provided
        elif chat_id not in Config.TELEGRAM_CHAT_ID:
            logger.warning(f"Attempted to send file to unauthorized chat {chat_id}. Using provided chat_id: {chat_id}")
            # Still use the provided chat_id even if not in authorized list (fallback)
            # The caller should have already validated this
            
        caption_title = f"FSA {area} - {current_category} Trip Details for {month_year}\n"
        caption = f"{caption_title}Total Trips: {trip_count}"

        # Send text message
        url_msg = f"https://api.telegram.org/bot{Config.TELEGRAM_BOT_TOKEN}/sendMessage"
        msg_payload = {'chat_id': chat_id, 'text': caption, 'parse_mode': 'HTML'}
        resp1 = requests.post(url_msg, data=msg_payload)
        if resp1.status_code == 200:
            logger.info(f"Sent summary message for {area} - {current_category} ({month_year}) to Telegram")
        else:
            logger.error(f"Failed to send summary message for {area} - {current_category} ({month_year}): {resp1.text}")

        # Send file
        url_file = f"https://api.telegram.org/bot{Config.TELEGRAM_BOT_TOKEN}/sendDocument"
        with open(file_path, 'rb') as f:
            data = {'chat_id': chat_id, 'caption': caption_title}
            files = {'document': (os.path.basename(file_path), f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
            resp2 = requests.post(url_file, data=data, files=files, stream=True)
        if resp2.status_code == 200 and resp2.json().get('ok'):
            logger.info(f"Sent Excel file for {area} - {current_category} ({month_year}): {file_path}")
        else:
            logger.error(f"Failed to send Excel file for {area} - {current_category} ({month_year}): {resp2.text}")

    except requests.RequestException as e:
        logger.error(f"Error sending to Telegram for {area} - {current_category} ({month_year}): {str(e)}")
        raise


@retry(stop=stop_after_attempt(3), wait=wait_exponential(multiplier=1, min=4, max=10), retry=retry_if_exception_type(PyMongoError))
async def process_batch_aggregation(collection, start_time, end_time, logger, current_category):
    """Process a single batch of trip data for MongoDB aggregation."""
    try:
        pipeline = [{'$match': {'createdAt': {'$gte': start_time, '$lte': end_time}, 'category': current_category, 'status': 'COMPLETED'}}, {
            '$project': {'_id': 0, 'Trip_Id': '$referenceId', 'Vehicle_Number': '$vehicleNumber',
                'Trip_Start_Time': {'$dateToString': {'format': '%Y-%m-%d %H:%M:%S', 'date': '$startTime', 'timezone': 'Asia/Kolkata'}},
                'Trip_End_Time': {'$dateToString': {'format': '%Y-%m-%d %H:%M:%S', 'date': '$endTime', 'timezone': 'Asia/Kolkata'}}, 'Trip_Category': '$category',
                'Filling_Quantity': '$fillingQuantity', 'Card_Quantity': '$cardQuantity', 'Filling_Station_Id': '$fillingStationId',
                'Filling_Station_Name': '$fillingStationName', 'Trip_Status': '$status', 'Dispensed_Quantity': '$dispensedQuantity', 'CMC_Number': {
                    '$cond': {'if': {'$and': [{'$isArray': '$request.dispensePoints'}, {'$gt': [{'$size': '$request.dispensePoints'}, 0]}]},
                        'then': {'$arrayElemAt': ['$request.dispensePoints.cmcNumber', 0]}, 'else': None}}, 'Customer_Name': {
                    '$cond': {'if': {'$and': [{'$isArray': '$request.dispensePoints'}, {'$gt': [{'$size': '$request.dispensePoints'}, 0]}]},
                        'then': {'$arrayElemAt': ['$request.dispensePoints.customerName', 0]}, 'else': None}}, 'Customer_Address': {
                    '$cond': {'if': {'$and': [{'$isArray': '$request.dispensePoints'}, {'$gt': [{'$size': '$request.dispensePoints'}, 0]}]},
                        'then': {'$arrayElemAt': ['$request.dispensePoints.address', 0]}, 'else': None}}}}]
        results = await collection.aggregate(pipeline).to_list(None)
        logger.debug(f"Fetched {len(results)} documents for category {current_category} from {start_time.strftime('%Y-%m-%d')} to {end_time.strftime('%Y-%m-%d')}")
        return results
    except PyMongoError as e:
        logger.error(f"Error in batch aggregation for category {current_category}: {str(e)}")
        raise


@retry(stop=stop_after_attempt(3), wait=wait_exponential(multiplier=1, min=4, max=10), retry=retry_if_exception_type(PyMongoError))
async def fetch_network_group_data(collection, filling_station_ids, logger):
    """Fetch network group data for given filling station IDs."""
    try:
        pipeline = [{'$match': {'code': {'$in': filling_station_ids}}}, {'$project': {'Filling_Station_Id': '$code', 'Area': {'$arrayElemAt': [{
            '$map': {'input': {'$filter': {'input': '$properties', 'as': 'prop', 'cond': {'$eq': ['$$prop.propName', 'area_name']}}}, 'as': 'filteredProp',
                'in': '$$filteredProp.value'}}, 0]}, '_id': 0}}]
        results = await collection.aggregate(pipeline).to_list(None)
        logger.debug(f"Fetched {len(results)} network_group records for matching stations")
        return results
    except PyMongoError as e:
        logger.error(f"Error fetching network_group data: {e}")
        raise


async def fetch_trip_data_for_area(client: AsyncIOMotorClient, logger: logging.Logger, area: str, current_category: str, start_time: datetime, end_time: datetime):
    """Fetch trip data filtered by area for a specific time range."""
    trip_collection = client["filling-station-service"]["trip"]
    network_collection = client["infra"]["network_group"]

    try:
        # Split the month into daily intervals for batch processing
        date_intervals = []
        current = start_time
        while current < end_time:
            next_interval = min(current + timedelta(days=1), end_time)
            date_intervals.append((current, next_interval))
            current = next_interval

        all_data = []
        # Process in batches using MAX_WORKERS
        for i in range(0, len(date_intervals), Config.MAX_WORKERS):
            batch = date_intervals[i:i + Config.MAX_WORKERS]
            results = await asyncio.gather(*(process_batch_aggregation(trip_collection, s, e, logger, current_category) for s, e in batch))
            for res in results:
                all_data.extend(res)

        trip_df = pd.DataFrame(all_data)
        
        # Clean up all_data to free memory
        del all_data
        gc.collect()
        
        if trip_df.empty:
            logger.warning(f"No trip data retrieved for area {area}, category {current_category} from {start_time.strftime('%Y-%m-%d')} to {end_time.strftime('%Y-%m-%d')}")
            return pd.DataFrame()

        filling_station_ids = trip_df['Filling_Station_Id'].dropna().unique().tolist()
        if not filling_station_ids:
            logger.warning(f"No filling station IDs found for area {area}, category {current_category}")
            return pd.DataFrame()
        
        network_df = pd.DataFrame(await fetch_network_group_data(network_collection, filling_station_ids, logger))
        network_df = network_df.drop_duplicates(subset=['Area', 'Filling_Station_Id'], keep='first')

        trip_df = trip_df.merge(network_df, on='Filling_Station_Id', how='left')
        
        # Clean up network_df to free memory
        del network_df
        gc.collect()
        
        # Filter by area
        trip_df = trip_df[trip_df['Area'] == area]
        
        if trip_df.empty:
            logger.warning(f"No trip data found for area {area} after filtering")
            return pd.DataFrame()

        desired_columns = [
            'Trip_Id', 'Vehicle_Number', 'Trip_Category', 'Trip_Status', 'Trip_Start_Time', 'Trip_End_Time', 'Area',
            'Dispensed_Quantity', 'Filling_Station_Name', 'Filling_Station_Id', 'Filling_Quantity', 'Card_Quantity',
            'CMC_Number', 'Customer_Name', 'Customer_Address'
        ]
        non_null_columns = [col for col in desired_columns if col in trip_df.columns and not trip_df[col].isna().all()]
        trip_df = trip_df[non_null_columns]

        return trip_df

    except Exception as e:
        logger.error(f"Error fetching trip data for area {area}, category {current_category}: {str(e)}")
        raise


@retry(stop=stop_after_attempt(3), wait=wait_exponential(multiplier=1, min=4, max=10), retry=retry_if_exception_type(Exception))
def save_to_excel(trip_df: pd.DataFrame, current_dir: str, logger: logging.Logger, area: str, current_category: str, month_year: str) -> str:
    """Save data to Excel with formatting."""
    try:
        if trip_df.empty:
            logger.warning(f"No data to save to Excel for area {area}, category {current_category} ({month_year})")
            return None

        # Sanitize area name for filename
        area_sanitized = sanitize_filename(area)
        # Format: Area-1_MC_Jan_2022.xlsx
        base_filename = f"{area_sanitized}_{current_category}_{month_year}"
        excel_path = os.path.join(current_dir, f"{base_filename}.xlsx")
        counter = 1
        while os.path.exists(excel_path):
            excel_path = os.path.join(current_dir, f"{base_filename}-{counter}.xlsx")
            counter += 1

        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            trip_df.to_excel(writer, sheet_name='Trip_Details', index=False)

        wb = load_workbook(excel_path)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        center_alignment = Alignment(horizontal='center', vertical='center')

        def format_sheet(ws):
            ws.freeze_panes = 'A2'
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = center_alignment
            for col_cells in ws.columns:
                max_length = max((len(str(cell.value)) if cell.value is not None else 0) for cell in col_cells)
                ws.column_dimensions[col_cells[0].column_letter].width = max(max_length * 1.2, 8)

        format_sheet(wb['Trip_Details'])
        wb.save(excel_path)
        logger.info(f"Saved Excel file for area {area}, category {current_category} ({month_year}): {excel_path}")
        
        # Clean up workbook to free memory
        wb.close()
        del wb
        gc.collect()
        
        return excel_path

    except Exception as e:
        logger.error(f"Failed to save Excel file for area {area}, category {current_category} ({month_year}): {str(e)}")
        raise




async def process_query_on_demand(client: AsyncIOMotorClient, current_dir: str, logger: logging.Logger, 
                                   categories: list, areas: list, start_time: datetime, end_time: datetime, 
                                   chat_id: int) -> None:
    """Process a query on-demand for multiple categories and multiple areas, send Excel files."""
    try:
        categories_str = ", ".join(categories) if categories else "all"
        areas_str = ", ".join(areas) if areas else "all"
        logger.info(f"Processing on-demand query: Categories={categories_str}, Areas={areas_str}, Period={start_time.strftime('%Y-%m-%d')} to {end_time.strftime('%Y-%m-%d')}")
        
        # Determine which categories to process
        if not categories or (len(categories) == 1 and categories[0] == "all"):
            categories_to_process = Config.CATEGORIES
            all_categories = True
        else:
            categories_to_process = [c for c in categories if c in Config.CATEGORIES]
            all_categories = False
        
        if not categories_to_process:
            error_msg = f"No valid categories found. Available categories: {', '.join(Config.CATEGORIES)}"
            logger.warning(error_msg)
            await send_message_to_telegram(chat_id, error_msg, logger)
            return
        
        # Determine which areas to process
        if not areas or (len(areas) == 1 and areas[0] == "all"):
            areas_to_process = Config.AREAS
            all_areas = True
        else:
            # Filter to only valid areas from Config.AREAS
            areas_to_process = [a for a in areas if a in Config.AREAS]
            all_areas = False
        
        if not areas_to_process:
            error_msg = f"No valid areas found. Available areas: {len(Config.AREAS)} areas"
            logger.warning(error_msg)
            await send_message_to_telegram(chat_id, error_msg, logger)
            return
        
        # Format month/year for filename
        if start_time.year == end_time.year and start_time.month == end_time.month:
            month_year = start_time.strftime('%b_%Y')
        else:
            month_year = f"{start_time.strftime('%b_%Y')}_to_{end_time.strftime('%b_%Y')}"
        
        total_files = 0
        total_trips = 0
        
        # Process each area and category combination
        for area in areas_to_process:
            for category in categories_to_process:
                try:
                    # Fetch trip data
                    trip_df = await fetch_trip_data_for_area(client, logger, area, category, start_time, end_time)
                    
                    if trip_df.empty:
                        logger.warning(f"No trip data found for {area}, category {category} for the specified period.")
                        continue
                    
                    # Get trip count before deleting DataFrame
                    trip_count = len(trip_df)
                    
                    # Save to Excel
                    excel_path = save_to_excel(trip_df, current_dir, logger, area, category, month_year)
                    
                    # Clean up DataFrame to free memory
                    del trip_df
                    gc.collect()
                    
                    if not excel_path:
                        logger.error(f"Failed to create Excel file for {area}, category {category}.")
                        continue
                    
                    # Send to Telegram
                    total_trips += trip_count
                    await send_to_telegram(excel_path, logger, area, category, month_year, trip_count, chat_id)
                    total_files += 1
                    
                    # Clean up after sending
                    try:
                        if os.path.exists(excel_path):
                            os.remove(excel_path)  # Remove temporary file
                    except Exception as e:
                        logger.warning(f"Could not remove temporary file {excel_path}: {str(e)}")
                    
                    gc.collect()  # Collect garbage after processing each file
                    
                except Exception as e:
                    logger.error(f"Error processing {area}, category {category}: {str(e)}")
                    # Ensure cleanup even on error
                    gc.collect()
                    continue
        
        if total_files == 0:
            error_msg = f"No trip data found for areas {areas_str}, categories {categories_str} for the specified period."
            await send_message_to_telegram(chat_id, error_msg, logger)
        else:
            summary_msg = f"✅ Processed {total_files} file(s) with total {total_trips} trips\nAreas: {areas_str}\nCategories: {categories_str}"
            await send_message_to_telegram(chat_id, summary_msg, logger)
        
    except Exception as e:
        error_msg = f"Error processing query: {str(e)}"
        logger.error(error_msg)
        await send_message_to_telegram(chat_id, error_msg, logger)


async def send_message_to_telegram(chat_id: int, message: str, logger: logging.Logger) -> None:
    """Send a text message to Telegram."""
    try:
        # Use chat_id if authorized, otherwise log warning but still send (fallback)
        if chat_id not in Config.TELEGRAM_CHAT_ID:
            logger.warning(f"Attempted to send message to unauthorized chat {chat_id}. Allowed chats: {Config.TELEGRAM_CHAT_ID}")
            # Still send to provided chat_id (caller should have validated)
        
        url_msg = f"https://api.telegram.org/bot{Config.TELEGRAM_BOT_TOKEN}/sendMessage"
        msg_payload = {'chat_id': chat_id, 'text': message, 'parse_mode': 'HTML'}
        resp = requests.post(url_msg, data=msg_payload)
        if resp.status_code == 200:
            logger.info(f"Sent message to Telegram chat {chat_id}")
        else:
            logger.error(f"Failed to send message to Telegram: {resp.text}")
    except Exception as e:
        logger.error(f"Error sending message to Telegram: {str(e)}")


async def handle_query(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle user query with NLP parsing."""
    chat_id = update.message.chat_id
    user_id = update.message.from_user.id
    
    logger = logging.getLogger(__name__)
    
    # Check if chat_id is in allowed chat IDs from Config
    if chat_id not in Config.TELEGRAM_CHAT_ID:
        logger.warning(f"Ignoring message from unauthorized chat {chat_id}. Allowed chats: {Config.TELEGRAM_CHAT_ID}")
        return ConversationHandler.END
    
    # Handle bot mentions in groups
    query = update.message.text.strip()
    
    # Remove bot mention if present
    if update.message.entities:
        for entity in update.message.entities:
            if entity.type == "mention":
                mentioned_text = query[entity.offset:entity.offset + entity.length]
                query = query.replace(mentioned_text, "").strip()
    
    logger.info(f"Received query from user {user_id} in chat {chat_id}: {query}")
    
    try:
        # Parse query with NLP
        parsed = parse_query_with_nlp(query, logger)
        
        categories = parsed.get("categories", [])
        category = parsed.get("category")  # For backward compatibility
        all_categories = parsed.get("all_categories", False)
        areas = parsed.get("areas", [])
        area = parsed.get("area")  # For backward compatibility
        all_areas = parsed.get("all_areas", False)
        period_text = parsed.get("period")
        has_period = parsed.get("has_period", False)
        has_area = parsed.get("has_area", False)
        
        # Store in context for conversation flow
        context.user_data['categories'] = categories
        context.user_data['category'] = category
        context.user_data['all_categories'] = all_categories
        context.user_data['areas'] = areas
        context.user_data['area'] = area
        context.user_data['all_areas'] = all_areas
        context.user_data['period_text'] = period_text
        context.user_data['has_period'] = has_period
        context.user_data['has_area'] = has_area
        context.user_data['chat_id'] = chat_id
        
        # Check if categories are found
        if not categories and not all_categories:
            await update.message.reply_text(
                "I couldn't find the trip category in your query. "
                f"Please specify one or more of: {', '.join(Config.CATEGORIES)}\n"
                "You can also say 'all categories' or 'all trips'.\n\n"
                "Examples:\n"
                "• 'Give me Excel file for PS trips for Area -1 for Jan 2025'\n"
                "• 'PS and MC trips Area 1 Jun 2024'\n"
                "• 'All categories Area 1 Jun 2024 to Aug 2024'\n"
                "• 'August trips' (all categories, last August)"
            )
            return ConversationHandler.END
        
        # If missing period, ask for it
        if not has_period:
            context.user_data['waiting_for'] = 'period'
            categories_display = "All categories" if all_categories else (", ".join(categories) if categories else "Unknown")
            await update.message.reply_text(
                f"Got it! Categories: {categories_display}\n"
                f"{f'Area: {area}' if area else 'Area: Not specified'}\n\n"
                "❓ For what period would you like the Excel file?\n"
                "Please provide:\n"
                "• Month and year (e.g., 'Jan 2025', 'January 2025')\n"
                "• Date range (e.g., 'Jun 2024 to Aug 2024')\n"
                "• Year only (e.g., '2025')\n"
                "• Month only (e.g., 'August' - will use last occurrence)"
            )
            return WAITING_FOR_PERIOD
        
        # If missing area, ask for it
        if not has_area:
            context.user_data['waiting_for'] = 'area'
            categories_display = "All categories" if all_categories else (", ".join(categories) if categories else "Unknown")
            area_list = "\n".join([f"• {area}" for area in Config.AREAS])
            await update.message.reply_text(
                f"Got it! Categories: {categories_display}\n"
                f"Period: {period_text}\n\n"
                "❓ For which area(s) would you like the Excel file?\n"
                f"Please specify:\n"
                "• Single area: 'Area-1', 'Area 1', or full name\n"
                "• Multiple areas: 'Area 1 and Area 2', 'Area-1, Area-2'\n"
                "• All areas: 'all areas' or 'all'\n\n"
                f"Available areas:\n{area_list}"
            )
            return WAITING_FOR_AREA
        
        # All information is present, process the query
        categories_display = "All categories" if all_categories else (", ".join(categories) if categories else "Unknown")
        areas_display = "All areas" if all_areas else (", ".join(areas) if areas else (area if area else "Unknown"))
        await update.message.reply_text(
            f"✅ Processing your request...\n\n"
            f"Categories: {categories_display}\n"
            f"Areas: {areas_display}\n"
            f"Period: {period_text}\n\n"
            "⏳ Please wait while I generate the Excel file(s)..."
        )
        
        # Parse the period to get start and end dates
        date_range = parse_date_from_text(period_text, logger)
        
        if not date_range:
            await update.message.reply_text(
                f"❌ Could not parse the period '{period_text}'. "
                "Please provide a valid date/period:\n"
                "• 'Jan 2025', 'January 2025'\n"
                "• 'Jun 2024 to Aug 2024'\n"
                "• '2025' (full year)\n"
                "• 'August' (month only - last occurrence)"
            )
            return ConversationHandler.END
        
        start_time, end_time = date_range
        
        # Process the query using connection manager
        current_dir = os.path.dirname(os.path.abspath(__file__))
        client = await _mongo_manager.get_client(logger)
        
        try:
            await process_query_on_demand(client, current_dir, logger, categories, areas, start_time, end_time, chat_id)
            # Last message time is already updated when get_client() was called
        finally:
            # Connection manager handles auto-close, but we can force cleanup if needed
            gc.collect()  # Collect garbage after query processing
        
        return ConversationHandler.END
        
    except Exception as e:
        logger.error(f"Error handling query: {str(e)}")
        await update.message.reply_text(
            f"❌ An error occurred while processing your query: {str(e)}\n\n"
            "Please try again or contact support."
        )
        return ConversationHandler.END


async def handle_period_response(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle period response from user."""
    chat_id = update.message.chat_id
    logger = logging.getLogger(__name__)
    
    # Check if chat_id is in allowed chat IDs from Config
    if chat_id not in Config.TELEGRAM_CHAT_ID:
        logger.warning(f"Ignoring message from unauthorized chat {chat_id}. Allowed chats: {Config.TELEGRAM_CHAT_ID}")
        return ConversationHandler.END
    
    # Update last message time for connection manager
    await _mongo_manager.get_client(logger)  # This updates last_message_time
    
    period_text = update.message.text.strip()
    context.user_data['period_text'] = period_text
    context.user_data['has_period'] = True
    
    categories = context.user_data.get('categories', [])
    all_categories = context.user_data.get('all_categories', False)
    area = context.user_data.get('area')
    has_area = context.user_data.get('has_area', False)
    
    # If area is missing, ask for it
    areas = context.user_data.get('areas', [])
    all_areas = context.user_data.get('all_areas', False)
    if not has_area or (not areas and not all_areas):
        context.user_data['waiting_for'] = 'area'
        categories_display = "All categories" if all_categories else (", ".join(categories) if categories else "Unknown")
        area_list = "\n".join([f"• {area}" for area in Config.AREAS])
        await update.message.reply_text(
            f"✅ Period: {period_text}\n"
            f"Categories: {categories_display}\n\n"
            "❓ For which area(s) would you like the Excel file?\n"
            f"Please specify:\n"
            "• Single area: 'Area-1', 'Area 1', or full name\n"
            "• Multiple areas: 'Area 1 and Area 2', 'Area-1, Area-2'\n"
            "• All areas: 'all areas' or 'all'\n\n"
            f"Available areas:\n{area_list}"
        )
        return WAITING_FOR_AREA
    
    # All information is present, process the query
    categories_display = "All categories" if all_categories else (", ".join(categories) if categories else "Unknown")
    areas_display = "All areas" if all_areas else (", ".join(areas) if areas else "Unknown")
    await update.message.reply_text(
        f"✅ Processing your request...\n\n"
        f"Categories: {categories_display}\n"
        f"Areas: {areas_display}\n"
        f"Period: {period_text}\n\n"
        "⏳ Please wait while I generate the Excel file(s)..."
    )
    
    # Parse the period to get start and end dates
    date_range = parse_date_from_text(period_text, logger)
    
    if not date_range:
        await update.message.reply_text(
            f"❌ Could not parse the period '{period_text}'. "
            "Please provide a valid date/period:\n"
            "• 'Jan 2025', 'January 2025'\n"
            "• 'Jun 2024 to Aug 2024'\n"
            "• '2025' (full year)\n"
            "• 'August' (month only - last occurrence)"
        )
        return ConversationHandler.END
    
    start_time, end_time = date_range
    chat_id = context.user_data.get('chat_id')
    
    # Process the query using connection manager
    current_dir = os.path.dirname(os.path.abspath(__file__))
    client = await _mongo_manager.get_client(logger)
    
    try:
        await process_query_on_demand(client, current_dir, logger, categories, areas, start_time, end_time, chat_id)
        # Last message time is already updated when get_client() was called
    finally:
        # Connection manager handles auto-close, but we can force cleanup if needed
        gc.collect()  # Collect garbage after query processing
    
    return ConversationHandler.END


async def handle_area_response(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle area response from user - supports multiple areas."""
    chat_id = update.message.chat_id
    logger = logging.getLogger(__name__)
    
    # Check if chat_id is in allowed chat IDs from Config
    if chat_id not in Config.TELEGRAM_CHAT_ID:
        logger.warning(f"Ignoring message from unauthorized chat {chat_id}. Allowed chats: {Config.TELEGRAM_CHAT_ID}")
        return ConversationHandler.END
    
    area_input = update.message.text.strip()
    
    # Check if user wants all areas
    if re.search(r'\ball\s+areas?\b|\ball\b', area_input, re.IGNORECASE):
        context.user_data['areas'] = ["all"]
        context.user_data['all_areas'] = True
        context.user_data['has_area'] = True
    else:
        # Parse multiple areas using NLP
        parsed = parse_query_with_nlp(area_input, logger)
        areas_found = parsed.get("areas", [])
        all_areas_flag = parsed.get("all_areas", False)
        
        if all_areas_flag or (areas_found and "all" in areas_found):
            context.user_data['areas'] = ["all"]
            context.user_data['all_areas'] = True
        elif areas_found:
            context.user_data['areas'] = areas_found
            context.user_data['all_areas'] = False
        else:
            # Try to match single or multiple areas manually
            areas_matched = []
            
            # Check for multiple area patterns
            area_numbers = re.findall(r'Area[- ]?(\d+)', area_input, re.IGNORECASE)
            if area_numbers:
                for num_str in area_numbers:
                    area_num = int(num_str)
                    if 1 <= area_num <= len(Config.AREAS):
                        area_matched = Config.AREAS[area_num - 1]
                        if area_matched not in areas_matched:
                            areas_matched.append(area_matched)
            
            # Also check for "and" pattern: "Area 1 and Area 2"
            if " and " in area_input.lower():
                parts = re.split(r'\s+and\s+', area_input, flags=re.IGNORECASE)
                for part in parts:
                    area_match = re.search(r'Area[- ]?(\d+)', part, re.IGNORECASE)
                    if area_match:
                        area_num = int(area_match.group(1))
                        if 1 <= area_num <= len(Config.AREAS):
                            area_matched = Config.AREAS[area_num - 1]
                            if area_matched not in areas_matched:
                                areas_matched.append(area_matched)
            
            # If no matches found, try single area matching
            if not areas_matched:
                for area in Config.AREAS:
                    if area_input.lower() in area.lower() or area.lower() in area_input.lower():
                        areas_matched.append(area)
                        break
            
            if areas_matched:
                context.user_data['areas'] = areas_matched
                context.user_data['all_areas'] = False
            else:
                await update.message.reply_text(
                    f"❌ Could not identify the area(s) '{area_input}'. "
                    "Please specify:\n"
                    "• Single area: 'Area-1', 'Area 1', or full name\n"
                    "• Multiple areas: 'Area 1 and Area 2', 'Area-1, Area-2'\n"
                    "• All areas: 'all areas' or 'all'"
                )
                return WAITING_FOR_AREA
        
        context.user_data['has_area'] = True
    
    categories = context.user_data.get('categories', [])
    all_categories = context.user_data.get('all_categories', False)
    areas = context.user_data.get('areas', [])
    all_areas = context.user_data.get('all_areas', False)
    period_text = context.user_data.get('period_text')
    
    # All information is present, process the query
    categories_display = "All categories" if all_categories else (", ".join(categories) if categories else "Unknown")
    areas_display = "All areas" if all_areas else (", ".join(areas) if areas else "Unknown")
    await update.message.reply_text(
        f"✅ Processing your request...\n\n"
        f"Categories: {categories_display}\n"
        f"Areas: {areas_display}\n"
        f"Period: {period_text}\n\n"
        "⏳ Please wait while I generate the Excel file(s)..."
    )
    
    # Parse the period to get start and end dates
    date_range = parse_date_from_text(period_text, logger)
    
    if not date_range:
        await update.message.reply_text(
            f"❌ Could not parse the period '{period_text}'. "
            "Please provide a valid date/period:\n"
            "• 'Jan 2025', 'January 2025'\n"
            "• 'Jun 2024 to Aug 2024'\n"
            "• '2025' (full year)\n"
            "• 'August' (month only - last occurrence)"
        )
        return ConversationHandler.END
    
    start_time, end_time = date_range
    chat_id = context.user_data.get('chat_id')
    
    # Process the query using connection manager
    current_dir = os.path.dirname(os.path.abspath(__file__))
    client = await _mongo_manager.get_client(logger)
    
    try:
        await process_query_on_demand(client, current_dir, logger, categories, areas, start_time, end_time, chat_id)
        # Last message time is already updated when get_client() was called
    finally:
        # Connection manager handles auto-close, but we can force cleanup if needed
        gc.collect()  # Collect garbage after query processing
    
    return ConversationHandler.END


async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Cancel the conversation."""
    chat_id = update.message.chat_id
    logger = logging.getLogger(__name__)
    
    # Check if chat_id is in allowed chat IDs from Config
    if chat_id not in Config.TELEGRAM_CHAT_ID:
        logger.warning(f"Ignoring cancel command from unauthorized chat {chat_id}. Allowed chats: {Config.TELEGRAM_CHAT_ID}")
        return ConversationHandler.END
    
    await update.message.reply_text("Operation cancelled.")
    context.user_data.clear()
    return ConversationHandler.END


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Start command handler."""
    chat_id = update.message.chat_id
    logger = logging.getLogger(__name__)
    
    # Check if chat_id is in allowed chat IDs from Config
    if chat_id not in Config.TELEGRAM_CHAT_ID:
        logger.warning(f"Ignoring /start command from unauthorized chat {chat_id}. Allowed chats: {Config.TELEGRAM_CHAT_ID}")
        return
    
    welcome_message = (
        "👋 Welcome to FSA Trip Data Bot!\n\n"
        "I can generate Excel files for trip data. Just ask me naturally!\n\n"
        "📝 Examples:\n"
        "• 'Give me Excel file for PS trips for Area -1 for Jun 2024'\n"
        "• 'PS and MC trips Area 1 Jun 2024 to Aug 2024'\n"
        "• 'All categories Area 1 Jun 2024'\n"
        "• 'August trips' (all categories, last August)\n"
        "• 'MC trips Area 5 for June 2023'\n\n"
        "✨ New Features:\n"
        "✅ Multiple categories: 'PS and MC trips'\n"
        "✅ Multiple areas: 'Area 1 and Area 2'\n"
        "✅ All areas: 'all areas'\n"
        "✅ Date ranges: 'Jun 2024 to Aug 2024'\n"
        "✅ Month-only: 'August' (finds last occurrence)\n"
        "✅ All categories: 'all categories' or 'all trips'\n\n"
        "Available Categories: " + ", ".join(Config.CATEGORIES) + "\n\n"
        "💡 Tip: Tag me (@your_bot_username) in a group or send me a message directly!"
    )
    await update.message.reply_text(welcome_message)


async def error_handler(update: object, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handle errors."""
    logger = logging.getLogger(__name__)
    logger.error(f"Update {update} caused error {context.error}")


async def run_bot() -> None:
    """Run the Telegram bot with NLP capabilities."""
    current_dir = os.path.dirname(os.path.abspath(__file__))
    logger = setup_logger(current_dir)
    logger.info("Starting Telegram Bot with NLP capabilities")
    
    # Check LLM7.io API key
    if not Config.LLM7_API_KEY:
        logger.warning("LLM7_API_KEY not set! NLP features will not work properly.")
        logger.warning("Please set LLM7_API_KEY environment variable with your llm7.io API key.")
        logger.info(f"Using LLM7.io endpoint: {Config.LLM7_BASE_URL}")
        logger.info(f"Using model: {Config.LLM7_MODEL}")
    else:
        logger.info(f"LLM7.io API configured with model: {Config.LLM7_MODEL}")
        logger.info(f"Using endpoint: {Config.LLM7_BASE_URL}")
    
    # Create application
    application = Application.builder().token(Config.TELEGRAM_BOT_TOKEN).build()
    
    # Handle bot mentions in groups
    async def handle_mention(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
        """Handle bot mentions in groups."""
        if update.message and update.message.text:
            chat_id = update.message.chat_id
            
            # Check if chat_id is in allowed chat IDs from Config
            if chat_id not in Config.TELEGRAM_CHAT_ID:
                logger.warning(f"Ignoring mention from unauthorized chat {chat_id}. Allowed chats: {Config.TELEGRAM_CHAT_ID}")
                return ConversationHandler.END
            
            # Check if bot is mentioned
            try:
                # Get bot info
                bot_info = await application.bot.get_me()
                bot_username = bot_info.username
                
                # Check if bot is mentioned in text or entities
                text_mentions = [e for e in (update.message.entities or []) if e.type == "mention"]
                text_user_mentions = [e for e in (update.message.entities or []) if e.type == "text_mention"]
                
                is_mentioned = False
                if text_mentions:
                    for entity in text_mentions:
                        mentioned = update.message.text[entity.offset:entity.offset + entity.length]
                        if mentioned == f"@{bot_username}":
                            is_mentioned = True
                            break
                
                if not is_mentioned and text_user_mentions:
                    is_mentioned = any(e.user.id == bot_info.id for e in text_user_mentions)
                
                if is_mentioned:
                    return await handle_query(update, context)
            except Exception as e:
                logger.error(f"Error checking bot mention: {str(e)}")
        
        return ConversationHandler.END
    
    # Add conversation handler for conversation states
    conv_handler = ConversationHandler(
        entry_points=[],
        states={
            WAITING_FOR_PERIOD: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_period_response)],
            WAITING_FOR_AREA: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_area_response)],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )
    
    # Add handlers - order matters! Add conversation handler first, then specific handlers
    application.add_handler(CommandHandler("start", start))
    application.add_handler(conv_handler)
    
    # Add mention handler for groups (must be after conversation handler)
    mention_handler = MessageHandler(
        filters.TEXT & filters.ChatType.GROUPS,
        handle_mention
    )
    application.add_handler(mention_handler)
    
    # Also handle direct messages (not in groups)
    direct_msg_handler = MessageHandler(
        filters.TEXT & filters.ChatType.PRIVATE & ~filters.COMMAND,
        handle_query
    )
    application.add_handler(direct_msg_handler)
    
    application.add_error_handler(error_handler)
    
    # Start the bot
    logger.info("Bot is running...")
    application.run_polling(allowed_updates=Update.ALL_TYPES)


async def main():
    """Main execution function - runs Telegram bot with NLP capabilities."""
    try:
        # Run the Telegram bot
        await run_bot()
    finally:
        # Cleanup on shutdown
        logger = logging.getLogger(__name__)
        await _mongo_manager.close(force=True)
        gc.collect()  # Final garbage collection
        logger.info("Application shutdown complete")


if __name__ == '__main__':
    # Validate configuration before starting
    if not Config.validate():
        print("\n⚠️  Configuration validation failed. Please set required environment variables.")
        sys.exit(1)
    
    asyncio.run(main())

